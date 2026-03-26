import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Fetch BBC News page
url = "https://www.bbc.com/news"
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
response = requests.get(url, headers=headers, timeout=15)
response.raise_for_status()

# Parse HTML
soup = BeautifulSoup(response.text, "html.parser")

# Collect news articles
articles = []
seen = set()

# Find news article links
for a_tag in soup.find_all("a", href=True):
    href = a_tag["href"]

    # Filter only news article URLs
    if "/news/articles/" not in href and "/news/videos/" not in href:
        continue

    # Build full URL
    if href.startswith("/"):
        href = "https://www.bbc.com" + href

    # Find headline text
    title = ""
    h_tag = a_tag.find(["h1", "h2", "h3", "h4", "h5", "h6"])
    if h_tag:
        title = h_tag.get_text(strip=True)
    else:
        # Look for title in span or paragraph tags
        span = a_tag.find("span") or a_tag.find("p")
        if span:
            title = span.get_text(strip=True)
        else:
            title = a_tag.get_text(strip=True)

    if not title or len(title) < 10:
        continue

    # Skip duplicates
    if title in seen:
        continue
    seen.add(title)

    # Search for date in parent elements
    date = ""
    parent = a_tag.parent
    for _ in range(5):
        if parent is None:
            break
        time_tag = parent.find("time")
        if time_tag:
            date = time_tag.get_text(strip=True) or time_tag.get("datetime", "")
            break
        parent = parent.parent

    articles.append({"title": title, "link": href, "date": date})

print(f"Found {len(articles)} articles.\n")
for i, a in enumerate(articles[:5], 1):
    print(f"{i}. {a['title']}")
    print(f"   {a['link']}")
    print(f"   Date: {a['date'] or 'Not found'}\n")

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "BBC News"

# Header styles
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2C3E50")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Write header row
headers_row = ["No", "Headline", "Link", "Date", "Scraped At"]
for col, header in enumerate(headers_row, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data styles
data_font = Font(name="Arial", size=10)
link_font = Font(name="Arial", size=10, color="1565C0", underline="single")
wrap_align = Alignment(vertical="top", wrap_text=True)
scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M")

# Write article data
for i, article in enumerate(articles, 1):
    row = i + 1

    ws.cell(row=row, column=1, value=i).font = data_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).border = thin_border

    ws.cell(row=row, column=2, value=article["title"]).font = data_font
    ws.cell(row=row, column=2).alignment = wrap_align
    ws.cell(row=row, column=2).border = thin_border

    ws.cell(row=row, column=3, value=article["link"]).font = link_font
    ws.cell(row=row, column=3).alignment = wrap_align
    ws.cell(row=row, column=3).border = thin_border

    ws.cell(row=row, column=4, value=article["date"] if article["date"] else "Not found").font = data_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).border = thin_border

    ws.cell(row=row, column=5, value=scraped_at).font = data_font
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5).border = thin_border

# Set column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20

# Add auto-filter and set header row height
ws.auto_filter.ref = f"A1:E{len(articles) + 1}"
ws.row_dimensions[1].height = 25

# Save output
output_path = "bbc_news.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")