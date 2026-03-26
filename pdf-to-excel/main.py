import os
import sys
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ─── Configuration ───
FOLDER_PATH = sys.argv[1] if len(sys.argv) > 1 else "."
OUTPUT_FILE = "pdf_contents.xlsx"
MAX_CELL_LENGTH = 32000  # Excel cell character limit is 32,767

# ─── Scan folder for PDF files ───
pdf_files = sorted([f for f in os.listdir(FOLDER_PATH) if f.lower().endswith(".pdf")])

if not pdf_files:
    print(f"No PDF files found in: {os.path.abspath(FOLDER_PATH)}")
    sys.exit(1)

print(f"Found {len(pdf_files)} PDF file(s) in '{os.path.abspath(FOLDER_PATH)}'.\n")

# ─── Extract text from each PDF ───
results = []

for filename in pdf_files:
    filepath = os.path.join(FOLDER_PATH, filename)
    file_size = os.path.getsize(filepath) / 1024  # KB

    try:
        reader = PdfReader(filepath)
        page_count = len(reader.pages)

        # Extract text from all pages
        full_text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"

        full_text = full_text.strip()

        # Truncate if exceeds Excel cell limit
        truncated = False
        if len(full_text) > MAX_CELL_LENGTH:
            full_text = full_text[:MAX_CELL_LENGTH] + "... [TRUNCATED]"
            truncated = True

        status = "Success"
        if not full_text:
            status = "No text found (possibly scanned)"
        elif truncated:
            status = "Truncated (exceeded cell limit)"

    except Exception as e:
        full_text = ""
        page_count = 0
        status = f"Error: {str(e)}"

    results.append({
        "filename": filename,
        "pages": page_count,
        "size_kb": round(file_size, 1),
        "text": full_text,
        "status": status,
    })

    print(f"  [{status}] {filename} ({page_count} pages, {file_size:.1f} KB)")

# ─── Create Excel workbook ───
wb = Workbook()
ws = wb.active
ws.title = "PDF Contents"

# Styles
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1B4F72")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
data_font = Font(name="Arial", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)

# Write header row
headers = ["No", "Filename", "Pages", "Size (KB)", "Extracted Text", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data rows
for i, r in enumerate(results, 1):
    row = i + 1

    ws.cell(row=row, column=1, value=i).font = data_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).border = thin_border

    ws.cell(row=row, column=2, value=r["filename"]).font = data_font
    ws.cell(row=row, column=2).border = thin_border

    ws.cell(row=row, column=3, value=r["pages"]).font = data_font
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).border = thin_border

    ws.cell(row=row, column=4, value=r["size_kb"]).font = data_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).border = thin_border

    ws.cell(row=row, column=5, value=r["text"]).font = data_font
    ws.cell(row=row, column=5).alignment = wrap_align
    ws.cell(row=row, column=5).border = thin_border

    status_font = Font(name="Arial", size=10, color="006100" if r["status"] == "Success" else "9C0006")
    ws.cell(row=row, column=6, value=r["status"]).font = status_font
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=6).border = thin_border

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 80
ws.column_dimensions["F"].width = 22

# Auto-filter and header row height
ws.auto_filter.ref = f"A1:F{len(results) + 1}"
ws.row_dimensions[1].height = 25

# ─── Summary sheet ───
summary = wb.create_sheet("Summary")
summary_data = [
    ("Scanned Folder", os.path.abspath(FOLDER_PATH)),
    ("Total PDFs", len(results)),
    ("Successful", sum(1 for r in results if r["status"] == "Success")),
    ("Failed / Empty", sum(1 for r in results if r["status"] != "Success")),
    ("Total Pages", sum(r["pages"] for r in results)),
    ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
]
for row, (label, value) in enumerate(summary_data, 1):
    summary.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True)
    summary.cell(row=row, column=2, value=value).font = Font(name="Arial")
summary.column_dimensions["A"].width = 20
summary.column_dimensions["B"].width = 50

# Save
wb.save(OUTPUT_FILE)
print(f"\nDone! Excel file saved: {OUTPUT_FILE}")
print(f"  Total: {len(results)} PDFs, {sum(r['pages'] for r in results)} pages")